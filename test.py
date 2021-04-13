import openpyxl
from pathlib import Path
import numpy as np
xlsx_file = Path('all.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)
sheet = wb_obj.active
print(sheet.max_row, sheet.max_column)
sentences = []
labels = []
for row in sheet.iter_rows():
    sentences.append(row[5].value)
    labels.append(row[8].value-1)

from keras.models import load_model
from keras.preprocessing.text import Tokenizer
from keras.preprocessing.sequence import pad_sequences
nlp = load_model("model.bin")

test = sentences
tokenizer = Tokenizer()
tokenizer.fit_on_texts(sentences)
data_point = tokenizer.texts_to_sequences(sentences)
y = np.array(labels)
vocab_size = len(tokenizer.word_index) 
print("Vocab:",vocab_size)
print()

test_doc = pad_sequences(tokenizer.texts_to_sequences(test), padding='post', maxlen=400)
with open("result.txt","w") as f:
    for p in nlp.predict_classes(test_doc):
        f.write(str(p+1)+"\n")