import openpyxl
from keras.preprocessing.text import Tokenizer
from keras.preprocessing.sequence import pad_sequences
from sklearn.model_selection import train_test_split
from keras.models import Sequential
from keras import layers
import numpy as np
from pathlib import Path
xlsx_file = Path('all.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)
sheet = wb_obj.active
print(sheet.max_row, sheet.max_column)
sentences = []
labels = []
for row in sheet.iter_rows():
    sentences.append(row[5].value)
    labels.append(row[8].value-1)

test_sentences = sentences

tokenizer = Tokenizer()
tokenizer.fit_on_texts(sentences)
data_point = tokenizer.texts_to_sequences(sentences)
y = np.array(labels)
vocab_size = len(tokenizer.word_index) 
print("Vocab:",vocab_size)
print()

maxlen = 400
embedding_dim = 400
data_point = pad_sequences(data_point, padding='post', maxlen=maxlen)

X_train, X_test, y_train, y_test = train_test_split(data_point, y, test_size=0.2, random_state=10)
data_point.shape

model = Sequential(name="NLP")
model.add(layers.Embedding(input_dim=vocab_size+1, 
                            output_dim=embedding_dim, 
                            input_length=maxlen))
model.add(layers.Conv1D(200, 5, activation='relu'))
model.add(layers.GlobalAveragePooling1D())
model.add(layers.Dense(100, activation='relu'))
model.add(layers.Dense(5, activation='softmax'))
model.compile(optimizer='adam',
              loss='sparse_categorical_crossentropy',
              metrics=['accuracy'])
model.summary()

"""Train Model"""
history = model.fit(X_train, y_train, epochs=100, validation_data=(X_test, y_test), batch_size=1000, verbose=1)
_,acc = history.model.evaluate(X_train, y_train, verbose=0)
print("Training Accuracy: {:.2f}%".format(acc*100))
_,acc = history.model.evaluate(X_test, y_test, verbose=0)
print("Testing Accuracy: {:.2f}%".format(acc*100))

from sklearn.metrics import f1_score, confusion_matrix
from keras.models import load_model
history.model.save("model.bin")
y_pred = history.model.predict_classes(X_test)
print(" F1 Score:\n",f1_score(y_test, y_pred, average="micro"))
print(" \n Confusion_Matrix:\n", confusion_matrix(y_test, y_pred))
